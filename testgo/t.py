import json
import openpyxl
import mysql.connector

exclude = [527444, 674736, 650905, 580674, 21514, 509185, 370500, 675157, 412208, 394111, 479785, 637305, 587526, 648910, 660495, 665756, 277940, 443447, 681322, 532030, 570017, 635860, 569225, 181383, 653623, 553175, 32447, 563363, 26614, 658199, 634324, 529606, 679762, 352167, 383583, 518992, 206537, 309807, 650927, 491129, 610445, 602033, 681577, 423926, 350382, 523582, 623789, 605837, 525705, 633246, 587339, 431065, 657016, 663992, 422739, 652707, 280284, 119029, 639616, 146102, 216949, 365821, 456084, 651304, 651364, 385375, 31579, 477543, 557838, 619115, 628816, 665414, 202580, 661044, 242038, 226057, 524117, 442629, 313852, 605432, 589436, 216941, 444772, 670047, 171626, 594308, 294336, 527451, 527860, 424557, 531304, 651504, 662071, 378524, 497858, 633706, 259148, 518417, 371476, 555252, 154913, 127916, 658065, 552210, 555766, 442672, 200034, 528207, 383088, 630199, 580559, 648233, 642825, 546811, 14265, 651025, 489666, 157871, 634573, 317976, 465165, 561899, 415713, 459547, 612725, 605667, 83359, 227907, 252788, 657946, 670473, 598399, 487519, 554856, 587753, 169537, 560362, 613893, 87470, 626641, 240074, 441492, 196741, 677196, 636340, 247490, 647437, 660827, 654707, 447355, 642785, 567383, 653923, 660210, 552367, 89596, 675454, 587817, 683350, 270319, 530320, 476513, 573704, 462441, 677299, 650138, 507853, 509173, 400526, 101247, 271851, 681483, 499822, 156602, 629879, 250663, 586585, 572287, 573918, 524524, 520702, 415068, 648536, 580077, 17170, 357659, 652493, 539975, 653653, 571617, 620441, 646525, 599065, 563023, 213902, 609268, 495887, 620152, 401216, 265875, 579993, 287731, 503641, 249386, 320129,
           411874, 526850, 640591, 498846, 522928, 649189, 677545, 660144, 544448, 200804, 537460, 414548, 488153, 656580, 339791, 502722, 510119, 643122, 654259, 281817, 405291, 586740, 648813, 332257, 574086, 546959, 677647, 510018, 683690, 244328, 660990, 522652, 652747, 647348, 440211, 347013, 607172, 658283, 592813, 422197, 635236, 251746, 656503, 615650, 492730, 447038, 437955, 601597, 577471, 283982, 654001, 244141, 650887, 121554, 681218, 677747, 681247, 660645, 665607, 562684, 448813, 529221, 455522, 531496, 650741, 520385, 654126, 648373, 615804, 206752, 485133, 516870, 513595, 19700, 495532, 402061, 491429, 651240, 183225, 438284, 381671, 429273, 448460, 605909, 676166, 624138, 347588, 121192, 367183, 568762, 658524, 549758, 658649, 677875, 659560, 563733, 121201, 662351, 529115, 640993, 662105, 647402, 314245, 554505, 408486, 488512, 480287, 479241, 639677, 636430, 531716, 568554, 472286, 631475, 432599, 148503, 648072, 664768, 670869, 177531, 506199, 610757, 496627, 679002, 305718, 593155, 427706, 390659, 390649, 681223, 523567, 651287, 659410, 650982, 463346, 677120, 672305, 415712, 658712, 260078, 380394, 460891, 606465, 666960, 448357, 597952, 582673, 201026, 644356, 573718, 296706, 335801, 658128, 641228, 387278, 199388, 638774, 573079, 532389, 239364, 458644, 632417, 454288, 643060, 154497, 663745, 39330, 607437, 600928, 505987, 464998, 356828, 648807, 402614, 578463, 680966, 644290, 592993, 679640, 371822, 567049, 655501, 638096, 650997, 196002, 531370, 681135, 421593, 126291, 632965, 560837, 252680, 550132, 674689, 383821, 604458, 594349, 418966, 577953, 579818, 535186, 658702, 581578, 676137, 527346]


bind = {527444: 'yuwarobyo@send4.uk', 674736: 'zhpliang80scn@163.com', 650905: 'rigubyohu@honeys.be', 580674: 'as83cf@bma.biglobe.ne.jp', 21514: 'fuyacho.honda@icloud.com', 509185: 'jh6rgdw4bg@privaterelay.appleid.com', 370500: 'junjielingmu902@gmail.com', 675157: 'bu29@mail4.uk', 412208: '1720790689@qq.com', 394111: '568052567@qq.com', 479785: 'nzqmth7mnw@privaterelay.appleid.com', 637305: 'qymygchvtt@privaterelay.appleid.com', 587526: 'poj27683@doolk.com', 648910: 'readystartsx1000@outlook.com', 660495: 'ywsj6p8466@privaterelay.appleid.com', 665756: 'kaori.t5445@gmail.com', 277940: 'y8xv9kcxjv@privaterelay.appleid.com', 443447: 'rk7h882mcv@privaterelay.appleid.com', 681322: 'hoshizuki@stayhome.li', 532030: 'seichandayo39@gmail.com', 570017: 'txmwtbxyv2@privaterelay.appleid.com', 635860: 'jiubaoyunren@gmail.com', 569225: 'hg3412675362@gmail.com', 181383: '7gbwx4bdzz@privaterelay.appleid.com', 653623: 'bpvmp2r59j@privaterelay.appleid.com', 553175: 's.sorat00128@gmail.com', 32447: 'kazamaakio@icloud.com', 563363: '1148681198@qq.com', 26614: 'krotenkopfsprung@gmail.com', 658199: 'vxiaoxiong@163.com', 634324: '3waco_o@ymail.ne.jp', 529606: 'coffee0402@i.softbank.jp', 679762: 'yuuuuumu.825@icloud.com', 352167: 'a0970212276@gmail.com', 383583: 'mark04100610@yahoo.ne.jp', 518992: 'cha.uni.top@icloud.com', 206537: 'csp4jwvjkg@privaterelay.appleid.com', 309807: 'xiaxingchen696@gmail.com', 650927: 'gone119@cocoro.uk', 491129: 'HFvery@gmail.com', 610445: 'qydtp6fqpv@privaterelay.appleid.com', 602033: '4nxtwhv9d2@privaterelay.appleid.com', 681577: 'objectfivetwilight@yahoo.co.jp', 423926: 'lionab1789@gmail.com', 350382: 'sakagty@gmail.com', 523582: '377548715@qq.com', 623789: 'jobbkawata@gmail.com', 605837: 'sw5yvfgtps@privaterelay.appleid.com', 525705: 'rs8t877pcf@privaterelay.appleid.com', 633246: 'koja0228@gmail.com', 587339: 'cfjsghfkz5@privaterelay.appleid.com', 431065: 'hr8qwkhhj4@privaterelay.appleid.com', 657016: 'wekfow@ritzw.com', 663992: 'vin9265cent@gmail.com', 422739: '5569221@gmail.com', 652707: 'agusyeri14@gmail.com', 280284: 'ns2bxdyrj2@privaterelay.appleid.com', 119029: 'combatto.7562386ssto@icloud.com', 639616: 'reishia014@yahoo.co.jp', 146102: 'gdozs17066@yahoo.co.jp', 216949: 'guoyuoting@gmail.com', 365821: 'danielteng20020605@gmail.com', 456084: '4skqvjggwv@privaterelay.appleid.com', 651304: 'zmvnvdsg29@privaterelay.appleid.com', 651364: '7dpb4sc6pc@privaterelay.appleid.com', 385375: 'a0960991190@gmail.com', 31579: 'pdbbrmjsc6@privaterelay.appleid.com', 477543: 'tzzdb7z8h5@privaterelay.appleid.com', 557838: 'bkzcnc44n6@privaterelay.appleid.com', 619115: 't.destiny.r@gmail.com', 628816: 'drbamboos@gmail.com', 665414: 'ttcz950@gmail.com', 202580: 'aa0981000126@gmail.com', 661044: 'fuka12121212@gmail.com', 242038: 'cgg5dt9ctq@privaterelay.appleid.com', 226057: 'jerry20020210@gmail.com', 524117: 'zeaplaygame02@gmail.com', 442629: 'iinetotemo114@icloud.com', 313852: 'smhmhzj4xw@privaterelay.appleid.com', 605432: 'm8wbkc9jwm@privaterelay.appleid.com', 589436: 'j.5@i.softbank.jp', 216941: 'w8stmcryp6@privaterelay.appleid.com', 444772: 'tkkytk@gmail.com', 670047: 'angelia1314@proton.me', 171626: '7xv8vmhk96@privaterelay.appleid.com', 594308: 'm0106k0305@icloud.com', 294336: 'qbynxwdp88@privaterelay.appleid.com', 527451: 'byota849@send4.uk', 527860: 'poka857@send4.uk', 424557: 'hinoamito2569.t@gmail.com', 531304: 'dqfvfzybgz@privaterelay.appleid.com', 651504: 'nhkssrv6kt@privaterelay.appleid.com', 662071: 'zvsix0326@icloud.com', 378524: 'br66j89pd4@privaterelay.appleid.com', 497858: 'tiyaki346@ruru.be', 633706: 'xxt0425xx@gmail.com', 259148: 'Michael800818@gmail.com', 518417: 'trwqh89hgn@privaterelay.appleid.com', 371476: '416416.teru@gmail.com', 555252: '79q8yvf4np@privaterelay.appleid.com', 154913: 'bb275999@gmail.com', 127916: 'fdv6gpx4zq@privaterelay.appleid.com', 658065: 't.oshima712@gmail.com', 552210: 'pogyusiga@sofia.re', 555766: '1440263326@qq.com', 442672: 'aries.pamo0412@gmail.com', 200034: 'g0922124330@gmail.com', 528207: 'ggxsp9hv7x@privaterelay.appleid.com', 383088: 'yasuchin1022@gmail.com', 630199: '1131888545@qq.com', 580559: '1492792066@qq.com', 648233: 'yyjhg970988@gmail.com', 642825: '820754676@qq.com', 546811: 'iradioiremember@163.com', 14265: 'satorunari0806@gmail.com', 651025: 'z7qtfnyggn@privaterelay.appleid.com', 489666: 't0983638525@gmail.com', 157871: 'qxcwx8tnqc@privaterelay.appleid.com', 634573: 'st45dc8bsv@privaterelay.appleid.com', 317976: 'eternalbraveheart0@gmail.com', 465165: '4sc97qhtfj@privaterelay.appleid.com', 561899: 'yvonne041223@gmail.com', 415713: 'leader4866@outlook.jp', 459547: '9f4wgh6q9p@privaterelay.appleid.com', 612725: 's5hvgbkr5f@privaterelay.appleid.com', 605667: 'tbv43470@ilebi.com', 83359: 'meeyan820@gmail.com', 227907: 'nemyuta@gmail.com', 252788: 'bn45353@gmail.com', 657946: 'xhdkyksbrh@privaterelay.appleid.com', 670473: '446200447@qq.com', 598399: 'saito.0820@icloud.com', 487519: 'nz6p572wgp@privaterelay.appleid.com', 554856: 'qnw5tyw7n8@privaterelay.appleid.com', 587753: 'makotin_530@icloud.com', 169537: 'eyshan1204@gmail.com', 560362: '8dwjncnbyg@privaterelay.appleid.com', 613893: 'k2fhz9xzng@privaterelay.appleid.com', 87470: 'fukuryu_sankonorei68@icloud.com', 626641: 'pobe2852@gmail.com', 240074: 'kagetorituka@gmail.com', 441492: 'pcykycdmm7@privaterelay.appleid.com', 196741: '907kaede@gmail.com', 677196: 'cc5jh69xkf@privaterelay.appleid.com', 636340: 'wowain@gmail.com', 247490: 'chopin067@gmail.com', 647437: 'wjgntwyqbd@privaterelay.appleid.com', 660827: '879kcxxm5c@privaterelay.appleid.com', 654707: 'g7mx6pg9kb@privaterelay.appleid.com', 447355: 'garo5604@gmail.com', 642785: 'kyomyo498@uma3.be', 567383: 'bg5gv4ynn4@privaterelay.appleid.com', 653923: 'cw5qq9z8br@privaterelay.appleid.com', 660210: 'navi9be@yahoo.co.jp', 552367: '848qsrbmrb@privaterelay.appleid.com', 89596: 'kirara0313kirakira@docomo.ne.jp', 675454: 'yiding.wang@ucdenver.edu', 587817: 'gkobwx@one-ml.com', 683350: 'q2ck5d4zrt@privaterelay.appleid.com', 270319: '1092073@gm.flhs.ptc.edu.tw', 530320: '410078487@qq.com', 476513: 'kotoh.26.mjiny@gmail.com', 573704: 'austinliu10303@gmail.com', 462441: '45556854@qq.com', 677299: '420563819@qq.com', 650138: 'nr8by2cvf9@privaterelay.appleid.com', 507853: 'masa.anko.1948@icloud.com', 509173: '75gqz9nvyc@privaterelay.appleid.com', 400526: 'baibaiskr@gmail.com', 101247: 'yu0614@gmail.com', 271851: '3525907164@qq.com', 681483: 'ckvota44120@gmail.com', 499822: 'z62mfwmvh2@privaterelay.appleid.com', 156602: 'mrb2wxrxbf@privaterelay.appleid.com', 629879: 'anraihonkeytonk@gmail.com', 250663: 'moyang51020@gmail.com', 586585: 'si6t.oota@rakumail.jp', 572287: '6pjsrjxzkb@privaterelay.appleid.com', 573918: '1577281425@qq.com', 524524: 'by.shigenoridd@icloud.com', 520702: 'qwer8811030405@icloud.com', 415068: 'jun0404jk@gmail.com', 648536: 'rc5mqwdrgb@privaterelay.appleid.com', 580077: 'jmbn79tdg6@privaterelay.appleid.com', 17170: 'dondonmaru920@gmail.com', 357659: 'a19961029@gmail.com', 652493: 'sxu21611@gmail.com', 539975: 'g9xwmm47jx@privaterelay.appleid.com', 653653: 'extrahi0613@gmail.com', 571617: 't2wcrbmqnp@privaterelay.appleid.com', 620441: 'tobiomaru@icloud.com', 646525: 'q6smkmpkqb@privaterelay.appleid.com', 599065: '806621313@qq.com', 563023: 'fgpcnrxsxw@privaterelay.appleid.com', 213902: 'jk52970605@gmail.com', 609268: 'hmiwa12121985@gmail.com', 495887: '42236486@qq.com', 620152: 'omobear0311@gmail.com', 401216: 'pnhfhgnfgf@privaterelay.appleid.com', 265875: 'jiantaixiaoyan979@gmail.com', 579993: 'nzdr86nd54@privaterelay.appleid.com', 287731: '5700ck@gmail.com', 503641: 't74tnknzqg@privaterelay.appleid.com', 249386: 'footaiob35@gmail.com', 320129: 'omobear0311@gmail.com', 411874: 'hiromu4228@gmail.com', 526850: 'kotoshimo100k@icloud.com', 640591: 'huchunhua@enjoypartytime.com', 498846: 'ke.ngood09@gmail.com', 522928: 'youqingxiaosen@gmail.com', 649189: '3040154741@QQ.COM', 677545: '574530821@qq.com', 660144: 'temporary@sika3.com', 544448: '5vhqtmdg4v@privaterelay.appleid.com', 200804: 'sakuraccho@gmail.com', 537460: 'wb4c4p7ybb@privaterelay.appleid.com', 414548: '874988765@qq.com', 488153: 'y_taguchi930@yahoo.co.jp', 656580: 'minisoba_3232@hotmail.co.jp', 339791: 's0936646292@gmail.com', 502722: '528bchvf7k@privaterelay.appleid.com', 510119: 'mcwr8nwgst@privaterelay.appleid.com', 643122: '1061328850@qq.com', 654259: '4f9vqmqbyb@privaterelay.appleid.com', 281817: 'x180188106@gmail.com', 405291: 'torubell3@gmail.com',
        586740: 'mujiang4872@foxmail.com', 648813: 'vd9s879tm6@privaterelay.appleid.com', 332257: 'rai.foo.0604@icloud.com', 574086: 't9fj2qrvj2@privaterelay.appleid.com', 546959: 'pdv5b2wrw8@privaterelay.appleid.com', 677647: 'shochin119@gmail.com', 510018: '937970667@qq.com', 683690: 'siyuan2687@gmail.com', 244328: '326007084@qq.com', 660990: 'weilee19960506@gmail.com', 522652: 'kinta2411@gmail.com', 652747: 'wh6nbqfgrq@privaterelay.appleid.com', 647348: 'yt7w5ybjjkd7z5xn9mb1@docomo.ne.jp', 440211: 'kvbg26z8bk@privaterelay.appleid.com', 347013: '18852308073@163.com', 607172: 'naka1214taka@icloud.com', 658283: '403302014@qq.com', 592813: '2kgp9crqfs@privaterelay.appleid.com', 422197: 'soya.splatoon.1220@gmail.com', 635236: 'a18072135@gmail.com', 251746: '4pnjbwyvmd@privaterelay.appleid.com', 656503: '469mv8qcnp@privaterelay.appleid.com', 615650: 'bekutorukkk@yahoo.co.jp', 492730: 'tsai13545@gmail.com', 447038: 'vq84stdt9r@privaterelay.appleid.com', 437955: 'pride-and-pathos.aki@ezweb.ne.jp', 601597: '4yt8ctbd5h@privaterelay.appleid.com', 577471: 'lkt091098@gmail.com', 283982: '9nn54ys64q@privaterelay.appleid.com', 654001: '8tgscv6k67@privaterelay.appleid.com', 244141: 'fighters0646@gmail.com', 650887: 'pyofoto@sika3.com', 121554: 'pvtm7w65z4@privaterelay.appleid.com', 681218: 'zupi849@hotsoup.be', 677747: 'zrkm7bytqz@privaterelay.appleid.com', 681247: 'andtheg@sute.jp', 660645: 'ywsj6p8466@privaterelay.appleid.com', 665607: 'xingyan157@gmail.com', 562684: 'laniidae.mz@gmail.com', 448813: 'zjj7ypkp6r@privaterelay.appleid.com', 529221: 'haru.221115@gmail.com', 455522: 'kansetsu200@i.softbank.jp', 531496: 'mugyafogo@f5.si', 650741: 'gvd8hrr849@privaterelay.appleid.com', 520385: 'tkjkb68jtp@privaterelay.appleid.com', 654126: 'wl02398507@gmail.com', 648373: 'a0930291131@gmail.com', 615804: 'jyari.imo.y1@gmail.com', 206752: 'bn12wyc@gmail.com', 485133: 'kaisei.0922@icloud.com', 516870: '2742113802@qq.com', 513595: 'yamasho8686@icloud.com', 19700: 'andou.rituko@gmail.com', 495532: 'fxk2gjgnhx@privaterelay.appleid.com', 402061: 'marxethic@gmail.com', 491429: 'vc8cyrprkm@privaterelay.appleid.com', 651240: 'sekimario@icloud.com', 183225: 'zebra870209@gmail.com', 438284: 'pphqzk7vwd@privaterelay.appleid.com', 381671: 'raul2520002000@yahoo.co.jp', 429273: 'arasan0108@icloud.com', 448460: '2ts7w64cvn@privaterelay.appleid.com', 605909: 'epj06143@ilebi.com', 676166: 'ben079505@gmail.com', 624138: 'junko329052@gmail.com', 347588: 'in.on.stop.push@gmail.com', 121192: 'premiumdog_vivi@yahoo.co.jp', 367183: 'khosoya0120@gmail.com', 568762: 'Vermillion@ezweb.ne.jp', 658524: 'rinsyuuten08@gmail.com', 549758: 'ymuta1111@gmail.comgipodo150@eay.jpbanu387@honeys.behuturyu@mbox.resapya420@exdonuts.comdaneka254@usako.netdoseyo488@instaddr.ukdamyozagyu@nekosan.ukpesakoto@digdig.orgpuzude@instaddr.ukyopiguze@ichigo.mej49cccjskf@privaterelay.appleid.comsilentsiren17@yahoo.co.jp', 658649: 'kesukuku.r@gmail.comkesukuku.dragon@gmail.com', 677875: 'ren.o_hara0427@icloud.com', 659560: 'japanwork2017@gmail.com', 563733: 'nrsdx2v4zy@privaterelay.appleid.com', 121201: 'sxnz54z7r4@privaterelay.appleid.com', 662351: '6m55zvzxb9@privaterelay.appleid.com', 529115: '1940103445@qq.com', 640993: '1019606443@qq.com437133533@qq.com', 662105: 'xuanyongc7@gmail.com', 647402: '0pb7374838p011b@ezweb.ne.jp', 314245: 'boss630825@gmail.comaoss630825@gmail.com', 554505: 'ffxc2mjvd4@privaterelay.appleid.com', 408486: 'cg9kb47dvq@privaterelay.appleid.com', 488512: 'zfhk9m2rf8@privaterelay.appleid.com', 480287: '2623871964@qq.com', 479241: 'mitutani@icloud.com', 639677: '2t7bc5rp9g@privaterelay.appleid.com', 636430: 'dy5zkmdm4m@privaterelay.appleid.com', 531716: 'hedf846fd7xht@gmail.com', 568554: 'hiroki.i7@docomo.ne.jp', 472286: 'gp6v7b27hd@privaterelay.appleid.com', 631475: 'kanaria9761@gmail.comRikutok2007@gmail.comAoi.tenshi401@gmail.comrikuto.kameyama0926@gmail.com', 432599: 'zen2naka@gmail.com', 148503: 'k8p7zj86t2@privaterelay.appleid.com', 648072: 'pk7kw576s2@privaterelay.appleid.com', 664768: 'y_kawanuma@yahoo.co.jp', 670869: 'kiwamomokawa7@gmail.comas83cf@bma.biglobe.ne.jp', 177531: 'badboys.forever1031@gmail.com', 506199: 'luna.2190.xyz@gmail.com', 610757: 'ryowa.7170@icloud.com', 496627: '6qb5jzb4qm@privaterelay.appleid.com', 679002: 'yessexy@icloud.com', 305718: 'masaru.sugiyama0114@gmail.com', 593155: '8dkxymtv9c@privaterelay.appleid.com', 427706: 'xuanming570@gmail.com', 390659: 'weqq3350@gmail.com', 390649: 'weqq3350@gmail.com', 681223: 'lubluis1510@gmail.comaVking1510@gmail.com', 523567: 'gsvctumo2525@gmail.com', 651287: 'cshi67960@gmail.comcshi67960@gmail.com', 659410: 'hehehehasiu911@gmail.comhogriderrrrr666@gmail.com', 650982: 'nmi308326@gmail.comnngg80679@gmail.com', 463346: 'shogaki.shogayaki@i.softbank.jp', 677120: 'noraafodor@gmail.com', 672305: 'masa4306@gmail.com', 415712: 'leader4866@outlook.jp', 658712: 'rinsyuuten08@gmail.com', 260078: 'shinoakira0725@gmail.com', 380394: 'hosaka073155@gmail.com', 460891: 'ryu918sei@icloud.com', 606465: 'toma.hayato.1223@icloud.com', 666960: 'box.mikumiku.3400abc@gmail.com', 448357: 'z48c846fbj@privaterelay.appleid.com', 597952: '8cpbz678d2@privaterelay.appleid.com', 582673: 'yuito.tokida@icloud.com', 201026: '578245827@qq.comruru1122q@gmail.comlol09139qwer@gmail.comkoalo903@gmail.comz65dgd959g@privaterelay.appleid.com', 644356: 'issyou_dou_desyo_usi_masu@docomo.ne.jp', 573718: 'katomo2019@gmail.com', 296706: 'tp.tsukurou@gmail.com', 335801: '2055449209@qq.com1044395803@qq.com', 658128: '1298577804@qq.com355636887@qq.com', 641228: '6ng4w5b9m8@privaterelay.appleid.com', 387278: 'arkrz@outlook.com', 199388: 'qlin66292@gmail.com15008243445@163.comsssss9965955@gmail.comnetmd@163.com1299031067@qq.com1253434138@qq.com2529304464@qq.com332282394@qq.comn0910747652@gmail.comwen79780@gmail.comq2011556612@yahoo.com.twyokai1710@gmail.com3586954808@qq.comeson2580@gmail.comg0922124330@gmail.comqwe8904312@gmail.comanimenokuso@gmail.comqdyqkcbxph@privaterelay.appleid.com7vjcdxcm7h@privaterelay.appleid.comAa091030271510@gmail.com840539793@qq.comph0996114@gmail.com', 638774: 'HFvery@gmail.com', 573079: 'kaikineri@gmail.com', 532389: 'sml209120@gmail.com', 239364: 'vovo12@ymail.ne.jp', 458644: '9sm2r89xdm@privaterelay.appleid.com', 632417: 'miyama382@gmail.com', 454288: 'fzgr8g22s8@privaterelay.appleid.com', 643060: 'kanisawa0109@gmail.com', 154497: 'apinkpanda801@gmail.com', 663745: '5knczknpmc@privaterelay.appleid.com', 39330: 'player.tide@gmail.com', 607437: 'wh5jt9c8wp@privaterelay.appleid.com', 600928: 'kfdhprj4zn@privaterelay.appleid.com', 505987: 'haruto200812@gmail.com', 464998: 'xqwk87gmg6@privaterelay.appleid.com', 356828: 'taigaheri@gmail.com', 648807: 'vd9s879tm6@privaterelay.appleid.comfxvkxy6z9z@privaterelay.appleid.com', 402614: 'azoth0972@gmail.com', 578463: 'takara.treasure0905@gmail.com', 680966: '273959012@qq.com', 644290: 'freely.holyhole.no.susume@gmail.comfreely.holyhole.no.susume@gmail.comfreely.holyhole.no.susume@gmail.com', 592993: 'chibe0178@gmail.com', 679640: 'toshi.s1041uma_uma@icloud.com', 371822: 'vrcz4gg2cr@privaterelay.appleid.com', 567049: 'shin.tosaka.922@gmail.comtosaka2344@gmail.com', 655501: 'k7rmnzhh65@privaterelay.appleid.com', 638096: 'neitiankaito321@gmail.comnoritama39@icloud.com', 650997: 'zaimassa1985130@gmail.com', 196002: 'kaminoge.city@gmail.com', 531370: '1106630497@qq.com', 681135: 'arsenal082468@gmail.com', 421593: 'kazuyoshi43@icloud.com', 126291: 'shigo910@icloud.com', 632965: 'xicunyicheng517@gmail.comkz.o2019.09@gmail.comm24013ok@takachiho.ac.jp', 560837: 'smh2634413060@gmail.com', 252680: 'aoisoccer20010402@gmail.com', 550132: 'uguug@qq.com384503788@qq.comlxc910614@163.comlxc9106142@126.com837006606@qq.com', 674689: 'sirius.cx330@gmail.com', 383821: '44mp6z2xw5@privaterelay.appleid.com', 604458: 'getukabizin@gmail.com', 594349: 'ahirugunsou@ymail.ne.jp', 418966: 'ooguroryuuya3627@gmail.com', 577953: '4pnw7hjw9v@privaterelay.appleid.com', 579818: '6t8ry45xtk@privaterelay.appleid.com', 535186: 'wall_e_robo@yahoo.co.jp', 658702: 'lidonghan833@gmail.com', 581578: 'sharurumarousu@gmail.com', 676137: 'euphonium.jam@gmail.com', 527346: '327141984@qq.com', 193901: '1515576469@qq.com', 606731: 'kattyanorixbuffaloes2022fan@gmail.com', 372440: '1715204638@qq.com', 552214: '2328914054@qq.com', 232529: 'fucchi@juno.ocn.ne.jp', 658019: 'mango830918@gmail.com', 249435: 'viaviasylvia@hotmail.com', 268141: 's0416644.kh@go.edu.tw', 648384: '3norixx@gmail.com'}


class Repo:
    host = '172.31.11.159'
    user = 'root'
    password = 'Mahjong2021'
    database = 'game_sync'

    def __init__(self):
        self.conn = mysql.connector.connect(
            host=self.host,
            user=self.user,
            password=self.password,
            database=self.database
        )

        self.conn.ping(reconnect=True)

    def get_cursor(self):
        return self.conn.cursor()

    def close(self):
        self.conn.close()

    def query_by_id(self, id: int):
        cursor = self.get_cursor()
        query = "SELECT * FROM platform WHERE database_id = %d" % id
        cursor.execute(query)

        results = cursor.fetchall()
        cursor.close()

        if len(results) == 0:
            return

        return results[0]

    def query_by_account(self, account_id):
        cursor = self.get_cursor()
        query = "SELECT * FROM platform WHERE account_id = %d" % account_id
        cursor.execute(query)

        results = cursor.fetchall()
        cursor.close()

        return results

    def query_by_address(self, addr: str):
        cursor = self.get_cursor()
        query = "SELECT * FROM platform WHERE address = '%s'" % addr
        cursor.execute(query)

        results = cursor.fetchall()
        cursor.close()

        return results

    def query_character_by_parent_id(self, p_id: int):
        cursor = self.get_cursor()
        query = "SELECT * FROM `character` WHERE parent_id = %d" % p_id
        cursor.execute(query)

        results = cursor.fetchall()
        cursor.close()

        return results

    def query_character_name_and_language_by_parent_id(self, p_id: int):
        cursor = self.get_cursor()
        query = "SELECT database_id, name, last_language FROM `character` WHERE parent_id = %d" % p_id
        cursor.execute(query)

        results = cursor.fetchall()
        cursor.close()

        return results


def check_by_account_id(r, id):
    p = r.query_by_id(id)
    if p is None:
        return

    ps = r.query_by_account(p[7])
    if len(ps) == 0:
        return

    for p in ps:
        if p[1] == 'Tourist' or p[1] == 'Steam':
            continue
        return p[6]


def check_by_address(repo: Repo, d_id: int):
    p = repo.query_by_id(d_id)
    if p is None:
        return

    ps = repo.query_by_address(p[5])
    if len(ps) == 0:
        return

    email = ''
    for p in ps:
        if p[6] != '':
            email += p[6]
    return email


def check_by_paipu_email(repo: Repo, d_id: int):
    p = repo.query_by_id(d_id)
    if p is None:
        return

    cs = repo.query_character_by_parent_id(p[7])
    if len(cs) == 0:
        return

    misc = cs[0][30]
    data = json.loads(misc)
    return data.get('306')


def get_id():
    # 打开 Excel 文件
    file_path = 'miss.xlsx'
    workbook = openpyxl.load_workbook(file_path)

    # 选择工作表
    sheet = workbook.active

    data = list(sheet.iter_rows(values_only=True))[1:]

    workbook.close()

    return list(filter(lambda row: row[0] not in exclude, data))


def get_all():
    # 打开 Excel 文件
    file_path = 'miss.xlsx'
    workbook = openpyxl.load_workbook(file_path)

    # 选择工作表
    sheet = workbook.active

    data = list(sheet.iter_rows(values_only=True))[1:]

    workbook.close()

    return data


if __name__ == '__main__':
    r = Repo()

    # 获取 id 列表
    ids = get_id()
    m = {}

    # account_id 关联查询数据
    # for d_id in ids:
    #     email = check_by_account_id(r, d_id[0])
    #     if email is not None and email != '':
    #         m[d_id[0]] = email

    # address 关联查询数据
    # for d_id in ids:
    #     email = check_by_address(r, d_id[0])
    #     if email is not None and email != '':
    #         m[d_id[0]] = email

    # paipu 关联查询数据
    # for d_id in ids:
    #     email = check_by_paipu_email(r, d_id[0])
    #     if email is not None and email != '':
    #         m[d_id[0]] = email

    # print(m)
    # print(m.keys())

    # # 写入新文件xlsx
    # workbook = openpyxl.Workbook()
    # sheet = workbook.active
    # sheet.append(['user_id', '金额', 'email'])

    # for v in get_all():
    #     b = bind.get(v[0], '')
    #     sheet.append([v[0], v[1], b])

    # file_path = 'found.xlsx'
    # workbook.save(file_path)

    # 通过 parent_id 查询 character 数据
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['user_id', 'role_id', 'email', 'name', 'language'])

    for k, v in bind.items():
        results = r.query_character_name_and_language_by_parent_id(k)
        result = results[0]
        sheet.append([k, result[0], v, result[1], result[2]])

    file_path = 'found_name_0321.xlsx'
    workbook.save(file_path)

    print("finish")
